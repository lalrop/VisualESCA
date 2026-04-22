"""
convert_data.py
---------------
Lee "Pre postulacion.xlsx" y genera "data.json" listo para ser
consumido por index.html en GitHub Pages.

Uso:
    python convert_data.py

Requisitos:
    pip install openpyxl
"""

import json
import math
import openpyxl
from pathlib import Path

EXCEL_FILE = Path(__file__).parent / "Pre postulacion.xlsx"
JSON_FILE  = Path(__file__).parent / "data.json"

COLUMN_MAP = {
    "Instrumento":                 "instrumento",
    "Codigo":                      "codigo",
    "Mes ejecucion":               "mes",
    "563_Base_Imponible":          "base_imponible",
    "Beneficio/perdida del periodo": "beneficio_perdida",
    "Ticket medio mensual":        "ticket_medio",
    "Margen operativo":            "margen_operativo",
}

def clean(value):
    """Convierte NaN/None a null para JSON."""
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    return value

def main():
    print(f"Leyendo {EXCEL_FILE.name}...")
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)
    ws = wb.active

    # Leer cabeceras de la primera fila
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    # Mapear índices de columna por nombre
    col_index = {}
    for excel_name, json_key in COLUMN_MAP.items():
        try:
            col_index[json_key] = headers.index(excel_name)
        except ValueError:
            print(f"  ADVERTENCIA: columna '{excel_name}' no encontrada en el Excel.")

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Saltar filas completamente vacías
        if not any(row):
            continue
        record = {key: clean(row[idx]) for key, idx in col_index.items()}
        # Saltar filas sin instrumento o codigo
        if not record.get("instrumento") or not record.get("codigo"):
            continue
        records.append(record)

    wb.close()

    with open(JSON_FILE, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, separators=(",", ":"))

    print(f"Generado {JSON_FILE.name} con {len(records)} registros.")

if __name__ == "__main__":
    main()
